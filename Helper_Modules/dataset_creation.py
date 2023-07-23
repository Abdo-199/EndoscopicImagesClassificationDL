from torchvision import transforms
from PIL import Image
from openpyxl import load_workbook
from pathlib import Path
import os
import shutil
import random

def apply_blur(image, kernel_size):
   blur =  transforms.GaussianBlur(kernel_size, (90, 100)) # high std
   transformed_image = blur(image)
   return transformed_image

def adjust_brightness(image, brightness_range):
  
    brightness_transform = transforms.ColorJitter(brightness=brightness_range)
    transformed_image = brightness_transform(image)
    return transformed_image

def mix_transforms(image, kernel_size, brightness_range):
    mix_transforms = transforms.Compose([transforms.GaussianBlur(kernel_size, (70, 90)),
                                         transforms.ColorJitter(brightness=brightness_range)])
    transformed_image = mix_transforms(image)
    return transformed_image

def apply_multi_transforms(images_list, brightness_range, magnitude_1, magnitude_2):
    """applies one of three transforms to each image a image list, to generate bad quality images\n
    transforms: brightness change, blur, and mix of blur and brightness change\n
    args:\n images_list: list of image paths\n
    brightness_range: int -> argument for transforms.ColorJitter(brightness_range)\n
    magnitude_1: Gaussian blur kernel size for only blur\n
    magnitude_1: Gaussian blur kernel size for the mix of blur and brightness change"""
    index = 0
    transformed_images = []
    # plt.figure(figsize=(18, 45))
    for image_file in images_list:

        image = Image.open(image_file)
        
        if index % 3 == 0:
            magnitude = magnitude_1
            magnitude = magnitude * 2 + 1    # to insure it will be odd
            transformed_image = apply_blur(image, magnitude)
        elif index % 3 == 1:
            transformed_image = adjust_brightness(image, brightness_range)
        elif index % 3 == 2:
            magnitude = magnitude_2
            magnitude = magnitude * 2 + 1
            transformed_image = mix_transforms(image, magnitude, brightness_range)

        transformed_images.append(transformed_image)
        index += 1
    return transformed_images

def find_cells_by_keyword(file_path, keywords:dict):
    """takes an excel file path and a dictionary of {cellIndex:"keyword"}
    returns the 0th cell of each row that satifies all the condistions in the dictionary"""
    result = []
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print("Error: Failed to load the Excel file.")
        print(str(e))
        return result
    for row in sheet.iter_rows(values_only=True):
        found_all_keywords = True

        for key, value in keywords.items():
            if value.lower() != str(row[key]).lower():
                found_all_keywords = False
                break

        if found_all_keywords:
            result.append(str(row[0]))

       
    workbook.close()
    return result

def create_quality_dataset(dataset_name, test_good, test_bad, train_good, train_bad_raw, train_bad_transformed, transformed_paths):
    test_good_path = os.path.join('data', dataset_name, 'test', 'good')
    os.makedirs(test_good_path)
    for path in test_good:
       shutil.copy(path, test_good_path)

    test_bad_path = os.path.join('data', dataset_name, 'test', 'bad')
    os.makedirs(test_bad_path)
    for path in test_bad:
       shutil.copy(path, test_bad_path)

    train_good_path = os.path.join('data', dataset_name, 'train', 'good')
    os.makedirs(train_good_path)
    for path in train_good:
       shutil.copy(path, train_good_path)

    train_bad_path = os.path.join('data', dataset_name, 'train', 'bad')
    os.makedirs(train_bad_path)
    for path in train_bad_raw:
       shutil.copy(path, train_bad_path)

    for path, img in zip(transformed_paths, train_bad_transformed):
      path = os.path.join(train_bad_path, Path(path).name)
      img.save(path)

def test_split_counter(directory:Path):
  
    class_dirs = [os.path.join(directory, name) for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]
    
    min_image_count = 100000
    for class_dir in class_dirs:
        image_count = len(list(Path(class_dir).glob("*\\*.png")))
        if image_count < min_image_count:
                min_image_count = image_count
        for patient_dir in os.listdir(class_dir):
            image_count = len(os.listdir(os.path.join(class_dir, patient_dir)))
            print(f"{image_count} in {patient_dir}")
        print("\n------------------------------------------\n")
    
    test_count = round(min_image_count/5)
    train_count = min_image_count - test_count
    print(f"min image in one class: {min_image_count}")
    print(f"select {test_count} images for the test dataset and delete it from the base dir. In case you did't do it")
    print(f"select {train_count} images for the test dataset and delete it from the base dir. In case you did't do it")
    return train_count, test_count

def train_dir(dataset_base_dir, dataset_dir, classes, equalize):
    
    min_image_count = 100000

    if equalize:
        for class_ in classes:
            image_count = len(list(Path(os.path.join(dataset_base_dir,class_)).glob("*\\*.png")))
            if image_count < min_image_count:
                    min_image_count = image_count

    for class_ in classes:
        all_src = list(Path(os.path.join(dataset_base_dir,class_)).glob("*\\*.png"))
        random.shuffle(all_src)
        index = 0
        for src_ in all_src:
            if index < min_image_count:
                shutil.copy(src_, os.path.join(dataset_dir, "train", class_, src_.name))
                index +=1
            else:
                break

def create_organ_dataset_1(dataset_name, test_ear, train_ear, test_nose, train_nose, test_vocal, train_vocal):
    """A function to create an organ dataset given the paths of all the classes and splits"""
    test_ear_path = os.path.join('data', dataset_name, 'test', 'ear')
    os.makedirs(test_ear_path)
    for path in test_ear:
       shutil.copy(path, test_ear_path)

    train_ear_path = os.path.join('data', dataset_name, 'train', 'ear')
    os.makedirs(train_ear_path)
    for path in train_ear:
       shutil.copy(path, train_ear_path)

    test_nose_path = os.path.join('data', dataset_name, 'test', 'nose')
    os.makedirs(test_nose_path)
    for path in test_nose:
       shutil.copy(path, test_nose_path)

    train_nose_path = os.path.join('data', dataset_name, 'train', 'nose')
    os.makedirs(train_nose_path)
    for path in train_nose:
       shutil.copy(path, train_nose_path)

    test_vocalfolds_path = os.path.join('data', dataset_name, 'test', 'vocalfolds')
    os.makedirs(test_vocalfolds_path)
    for path in test_vocal:
       shutil.copy(path, test_vocalfolds_path)

    train_vocal_path = os.path.join('data', dataset_name, 'train', 'vocalfolds')
    os.makedirs(train_vocal_path)
    for path in train_vocal:
       shutil.copy(path, train_vocal_path)

def create_organ_dataset_2(dataset_dir:Path, classes, original_dir:Path, equalize=True):
    """The second version of the dataset creation function 
    \nargs:\n
        dataset_dir: the Path() of the dataset to be created\n
        classes: the classes to be created\n
        original_dir: the Path() of the central database\n
        equalize: should the classes be equalized"""
    random.seed(42)

    for class_ in classes:
        os.makedirs(os.path.join(dataset_dir, "train", class_), exist_ok=True)
        os.makedirs(os.path.join(dataset_dir, "test", class_), exist_ok=True)
        
    base_dir = os.path.join("data", f"{dataset_dir.name}_base")

    shutil.copytree(original_dir, base_dir)
    
    trin_count, test_count = test_split_counter(original_dir)

    if not equalize:
        test_count = 100000
    
    print('inter the test directories for each organ splited by ","')
    test_ear = input('Ear')
    test_nose = input('Nose')
    test_vocal = input('vocal')

    ear_images = list()
    for dir in test_ear.split(","):
        ear_images.extend(list(Path(os.path.join(base_dir, "Ear", dir)).glob("*.png")))

    random.shuffle(ear_images)
    for image in ear_images:
        if ear_images.index(image) < test_count:
            shutil.move(image, os.path.join(dataset_dir, "test", "Ear"))
        else:
            os.remove(image)

    nose_images = list()
    for dir in test_nose.split(","):
        nose_images.extend(list(Path(os.path.join(base_dir, "Nose", dir)).glob("*.png")))
    random.shuffle(nose_images)
    for image in nose_images:
        if nose_images.index(image) < test_count:
            shutil.move(image, os.path.join(dataset_dir, "test", "Nose"))
        else:
            os.remove(image)

    vocal_images = list()
    for dir in test_vocal.split(","):
        vocal_images.extend(list(Path(os.path.join(base_dir, "VocalFolds", dir)).glob("*.png")))
    random.shuffle(vocal_images)
    for image in vocal_images:
        if vocal_images.index(image) < test_count:
            shutil.move(image, os.path.join(dataset_dir, "test", "VocalFolds"))
        else:
            os.remove(image)

    train_dir(base_dir, dataset_dir, classes, equalize)